"""
Excel处理模块
负责Excel文件的生成、导出和VBA模板处理
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import os


# 样式定义
HEADER_FONT = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F77B4', end_color='1F77B4', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

DATA_FONT = Font(name='微软雅黑', size=10)
DATA_ALIGNMENT = Alignment(horizontal='center', vertical='center')

NUMBER_FORMAT = '#,##0.00'
PERCENTAGE_FORMAT = '0.00%'

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

TITLE_FONT = Font(name='微软雅黑', bold=True, size=16, color='1F77B4')
SUBTITLE_FONT = Font(name='微软雅黑', bold=True, size=12, color='333333')


def apply_header_style(ws, row, col):
    """应用表头样式"""
    cell = ws.cell(row=row, column=col)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


def apply_data_style(ws, row, col, is_number=False):
    """应用数据样式"""
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGNMENT
    cell.border = THIN_BORDER
    
    if is_number:
        cell.number_format = NUMBER_FORMAT


def apply_title_style(ws, row, col, title):
    """应用标题样式"""
    cell = ws.cell(row=row, column=col)
    cell.value = title
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_raw_data_sheet(wb, df, sheet_name='原始数据'):
    """创建原始数据工作表"""
    ws = wb.create_sheet(sheet_name)
    
    # 写入标题
    apply_title_style(ws, 1, 1, '原始数据')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    
    # 写入表头
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=3, column=col_idx, value=col_name)
        apply_header_style(ws, 3, col_idx)
    
    # 写入数据
    for row_idx, row in enumerate(df.values, 4):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
            is_number = isinstance(value, (int, float))
            apply_data_style(ws, row_idx, col_idx, is_number)
    
    # 添加汇总行
    summary_row = len(df) + 5
    ws.cell(row=summary_row, column=1, value='汇总')
    ws.cell(row=summary_row, column=1).font = SUBTITLE_FONT
    
    numeric_cols = df.select_dtypes(include=['number']).columns
    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name in numeric_cols:
            ws.cell(row=summary_row, column=col_idx, value=df[col_name].sum())
            ws.cell(row=summary_row, column=col_idx).number_format = NUMBER_FORMAT
            ws.cell(row=summary_row, column=col_idx).font = SUBTITLE_FONT
    
    auto_adjust_column_width(ws)
    
    return ws


def create_summary_sheet(wb, df, sheet_name='统计汇总'):
    """创建统计汇总工作表"""
    ws = wb.create_sheet(sheet_name)
    
    # 标题
    apply_title_style(ws, 1, 1, '统计汇总报表')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    
    # 报表信息
    ws.cell(row=2, column=1, value=f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    ws.cell(row=2, column=1).font = Font(italic=True, color='666666')
    
    # 数值列统计
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
    if not numeric_cols:
        ws.cell(row=4, column=1, value='未找到数值列')
        return ws
    
    # 统计表头
    headers = ['指标', '计数', '均值', '标准差', '最小值', '25%', '50%', '75%', '最大值', '总和']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)
        apply_header_style(ws, 4, col_idx)
    
    # 统计数据
    for row_idx, col_name in enumerate(numeric_cols, 5):
        col_data = df[col_name]
        
        ws.cell(row=row_idx, column=1, value=col_name)
        ws.cell(row=row_idx, column=2, value=int(col_data.count()))
        ws.cell(row=row_idx, column=3, value=float(col_data.mean()))
        ws.cell(row=row_idx, column=4, value=float(col_data.std()))
        ws.cell(row=row_idx, column=5, value=float(col_data.min()))
        ws.cell(row=row_idx, column=6, value=float(col_data.quantile(0.25)))
        ws.cell(row=row_idx, column=7, value=float(col_data.median()))
        ws.cell(row=row_idx, column=8, value=float(col_data.quantile(0.75)))
        ws.cell(row=row_idx, column=9, value=float(col_data.max()))
        ws.cell(row=row_idx, column=10, value=float(col_data.sum()))
        
        # 应用样式
        for col_idx in range(1, 11):
            apply_data_style(ws, row_idx, col_idx, col_idx > 1)
    
    # 添加图表
    if len(numeric_cols) > 0:
        create_summary_chart(ws, len(numeric_cols), len(headers))
    
    auto_adjust_column_width(ws)
    
    return ws


def create_summary_chart(ws, data_rows, data_cols):
    """创建汇总图表"""
    # 柱状图
    chart = BarChart()
    chart.type = "col"
    chart.title = "各指标汇总"
    chart.y_axis.title = "金额"
    chart.x_axis.title = "指标"
    
    # 数据区域
    data = Reference(ws, min_col=10, min_row=4, max_row=4 + data_rows)
    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + data_rows)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 20
    chart.height = 12
    
    ws.add_chart(chart, "A" + str(data_rows + 7))


def create_category_summary_sheet(wb, df, category_col, amount_col, sheet_name='分类汇总'):
    """创建分类汇总工作表"""
    ws = wb.create_sheet(sheet_name)
    
    # 标题
    apply_title_style(ws, 1, 1, '分类汇总报表')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    
    # 分类汇总数据
    category_stats = df.groupby(category_col).agg(
        金额合计=(amount_col, 'sum'),
        金额均值=(amount_col, 'mean'),
        交易次数=(amount_col, 'count')
    ).reset_index()
    
    category_stats['占比'] = category_stats['金额合计'] / category_stats['金额合计'].sum()
    category_stats = category_stats.sort_values('金额合计', ascending=False)
    
    # 表头
    headers = [category_col, '金额合计', '金额均值', '交易次数', '占比']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=header)
        apply_header_style(ws, 3, col_idx)
    
    # 数据
    for row_idx, (_, row) in enumerate(category_stats.iterrows(), 4):
        ws.cell(row=row_idx, column=1, value=row[category_col])
        ws.cell(row=row_idx, column=2, value=row['金额合计'])
        ws.cell(row=row_idx, column=3, value=row['金额均值'])
        ws.cell(row=row_idx, column=4, value=int(row['交易次数']))
        ws.cell(row=row_idx, column=5, value=row['占比'])
        
        for col_idx in range(1, 6):
            apply_data_style(ws, row_idx, col_idx, col_idx > 1)
        
        ws.cell(row=row_idx, column=2).number_format = NUMBER_FORMAT
        ws.cell(row=row_idx, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=row_idx, column=5).number_format = PERCENTAGE_FORMAT
    
    # 饼图
    create_pie_chart(ws, len(category_stats), category_col)
    
    auto_adjust_column_width(ws)
    
    return ws


def create_pie_chart(ws, data_rows, category_name):
    """创建饼图"""
    chart = PieChart()
    chart.title = f"{category_name}占比分析"
    chart.width = 18
    chart.height = 12
    
    data = Reference(ws, min_col=2, min_row=3, max_row=3 + data_rows)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + data_rows)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A" + str(data_rows + 6))


def create_monthly_trend_sheet(wb, df, date_col, amount_col, sheet_name='月度趋势'):
    """创建月度趋势工作表"""
    ws = wb.create_sheet(sheet_name)
    
    # 标题
    apply_title_style(ws, 1, 1, '月度趋势报表')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    
    # 月度汇总
    df_copy = df.copy()
    df_copy[date_col] = pd.to_datetime(df_copy[date_col], errors='coerce')
    df_copy['年月'] = df_copy[date_col].dt.to_period('M')
    
    monthly = df_copy.groupby('年月').agg(
        金额合计=(amount_col, 'sum'),
        金额均值=(amount_col, 'mean'),
        交易次数=(amount_col, 'count')
    ).reset_index()
    
    monthly['年月'] = monthly['年月'].astype(str)
    monthly['环比增长'] = monthly['金额合计'].pct_change()
    
    # 表头
    headers = ['年月', '金额合计', '金额均值', '交易次数', '环比增长']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=header)
        apply_header_style(ws, 3, col_idx)
    
    # 数据
    for row_idx, (_, row) in enumerate(monthly.iterrows(), 4):
        ws.cell(row=row_idx, column=1, value=row['年月'])
        ws.cell(row=row_idx, column=2, value=row['金额合计'])
        ws.cell(row=row_idx, column=3, value=row['金额均值'])
        ws.cell(row=row_idx, column=4, value=int(row['交易次数']))
        ws.cell(row=row_idx, column=5, value=row['环比增长'])
        
        for col_idx in range(1, 6):
            apply_data_style(ws, row_idx, col_idx, col_idx > 1)
        
        ws.cell(row=row_idx, column=2).number_format = NUMBER_FORMAT
        ws.cell(row=row_idx, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=row_idx, column=5).number_format = PERCENTAGE_FORMAT
    
    # 折线图
    create_line_chart(ws, len(monthly))
    
    auto_adjust_column_width(ws)
    
    return ws


def create_line_chart(ws, data_rows):
    """创建折线图"""
    chart = LineChart()
    chart.title = "月度趋势"
    chart.y_axis.title = "金额"
    chart.x_axis.title = "月份"
    chart.width = 20
    chart.height = 12
    
    data = Reference(ws, min_col=2, min_row=3, max_row=3 + data_rows, max_col=2)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + data_rows)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A" + str(data_rows + 6))


def export_to_excel(df, filename, export_options, include_charts=True):
    """
    导出为Excel文件
    
    Args:
        df: 数据DataFrame
        filename: 文件名
        export_options: 导出内容选项列表
        include_charts: 是否包含图表
    
    Returns:
        BytesIO对象
    """
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 根据选项创建不同的工作表
    if '原始数据' in export_options:
        create_raw_data_sheet(wb, df)
    
    if '统计汇总' in export_options:
        create_summary_sheet(wb, df)
    
    if '分类汇总' in export_options:
        # 尝试找到分类列
        category_cols = df.select_dtypes(include=['object']).columns
        numeric_cols = df.select_dtypes(include=['number']).columns
        
        if len(category_cols) > 0 and len(numeric_cols) > 0:
            create_category_summary_sheet(
                wb, df, 
                category_cols[0], 
                numeric_cols[0]
            )
    
    if '数据透视表' in export_options:
        category_cols = df.select_dtypes(include=['object']).columns
        numeric_cols = df.select_dtypes(include=['number']).columns
        
        if len(category_cols) >= 2 and len(numeric_cols) > 0:
            create_pivot_table_sheet(
                wb, df,
                category_cols[0],
                category_cols[1],
                numeric_cols[0]
            )
    
    # 保存到BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def create_pivot_table_sheet(wb, df, index_col, columns_col, values_col, sheet_name='数据透视表'):
    """创建数据透视表工作表"""
    ws = wb.create_sheet(sheet_name)
    
    # 标题
    apply_title_style(ws, 1, 1, '数据透视表')
    
    # 创建透视表
    pivot = pd.pivot_table(
        df,
        values=values_col,
        index=index_col,
        columns=columns_col,
        aggfunc='sum',
        fill_value=0
    )
    
    # 表头
    ws.cell(row=3, column=1, value=index_col)
    apply_header_style(ws, 3, 1)
    
    for col_idx, col_name in enumerate(pivot.columns, 2):
        ws.cell(row=3, column=col_idx, value=col_name)
        apply_header_style(ws, 3, col_idx)
    
    # 数据
    for row_idx, (idx, row) in enumerate(pivot.iterrows(), 4):
        ws.cell(row=row_idx, column=1, value=idx)
        apply_data_style(ws, row_idx, 1)
        
        for col_idx, value in enumerate(row, 2):
            ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(ws, row_idx, col_idx, True)
    
    # 添加汇总行和列
    summary_row = len(pivot) + 4
    ws.cell(row=summary_row, column=1, value='合计')
    ws.cell(row=summary_row, column=1).font = SUBTITLE_FONT
    
    for col_idx in range(2, len(pivot.columns) + 2):
        total = sum(pivot.iloc[:, col_idx - 2])
        ws.cell(row=summary_row, column=col_idx, value=total)
        ws.cell(row=summary_row, column=col_idx).number_format = NUMBER_FORMAT
        ws.cell(row=summary_row, column=col_idx).font = SUBTITLE_FONT
    
    auto_adjust_column_width(ws)
    
    return ws


def export_with_vba_template(df, template_name):
    """
    使用VBA模板导出
    
    Args:
        df: 数据DataFrame
        template_name: 模板文件名
    
    Returns:
        BytesIO对象
    """
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 创建数据工作表
    ws_data = wb.create_sheet('财务数据')
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_data.cell(row=r_idx, column=c_idx, value=value)
    
    # 创建报表工作表
    ws_report = wb.create_sheet('财务报表')
    
    # 标题
    apply_title_style(ws_report, 1, 1, '财务综合报表')
    ws_report.merge_cells('A1:F1')
    
    # 基本信息
    ws_report.cell(row=2, column=1, value='报表日期:')
    ws_report.cell(row=2, column=2, value=datetime.now().strftime('%Y-%m-%d'))
    ws_report.cell(row=3, column=1, value='数据来源:')
    ws_report.cell(row=3, column=2, value='自动化导入')
    
    # 汇总数据
    numeric_cols = df.select_dtypes(include=['number']).columns
    if len(numeric_cols) > 0:
        ws_report.cell(row=5, column=1, value='科目')
        ws_report.cell(row=5, column=2, value='金额')
        ws_report.cell(row=5, column=3, value='占比')
        
        apply_header_style(ws_report, 5, 1)
        apply_header_style(ws_report, 5, 2)
        apply_header_style(ws_report, 5, 3)
        
        total = df[numeric_cols[0]].sum()
        for row_idx, col_name in enumerate(numeric_cols, 6):
            ws_report.cell(row=row_idx, column=1, value=col_name)
            ws_report.cell(row=row_idx, column=2, value=df[col_name].sum())
            ws_report.cell(row=row_idx, column=3, value=df[col_name].sum() / total if total != 0 else 0)
            
            apply_data_style(ws_report, row_idx, 1)
            apply_data_style(ws_report, row_idx, 2, True)
            apply_data_style(ws_report, row_idx, 3, True)
            
            ws_report.cell(row=row_idx, column=2).number_format = NUMBER_FORMAT
            ws_report.cell(row=row_idx, column=3).number_format = PERCENTAGE_FORMAT
    
    auto_adjust_column_width(ws_report)
    
    # 保存到BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def create_monthly_report(df, date_col, amount_col, category_col=None):
    """
    创建月度报表
    
    Args:
        df: 数据DataFrame
        date_col: 日期列
        amount_col: 金额列
        category_col: 分类列（可选）
    
    Returns:
        BytesIO对象
    """
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 数据准备
    df_copy = df.copy()
    df_copy[date_col] = pd.to_datetime(df_copy[date_col], errors='coerce')
    df_copy['年月'] = df_copy[date_col].dt.to_period('M')
    
    # 月度汇总
    monthly = df_copy.groupby('年月').agg(
        金额合计=(amount_col, 'sum'),
        金额均值=(amount_col, 'mean'),
        交易次数=(amount_col, 'count')
    ).reset_index()
    
    monthly['年月'] = monthly['年月'].astype(str)
    monthly['环比增长'] = monthly['金额合计'].pct_change()
    
    # 创建月度报表工作表
    ws = wb.create_sheet('月度报表')
    
    apply_title_style(ws, 1, 1, '月度财务报表')
    ws.merge_cells('A1:E1')
    
    # 表头
    headers = ['年月', '金额合计', '金额均值', '交易次数', '环比增长']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=header)
        apply_header_style(ws, 3, col_idx)
    
    # 数据
    for row_idx, (_, row) in enumerate(monthly.iterrows(), 4):
        ws.cell(row=row_idx, column=1, value=row['年月'])
        ws.cell(row=row_idx, column=2, value=row['金额合计'])
        ws.cell(row=row_idx, column=3, value=row['金额均值'])
        ws.cell(row=row_idx, column=4, value=int(row['交易次数']))
        ws.cell(row=row_idx, column=5, value=row['环比增长'])
        
        for col_idx in range(1, 6):
            apply_data_style(ws, row_idx, col_idx, col_idx > 1)
        
        ws.cell(row=row_idx, column=2).number_format = NUMBER_FORMAT
        ws.cell(row=row_idx, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=row_idx, column=5).number_format = PERCENTAGE_FORMAT
    
    # 折线图
    chart = LineChart()
    chart.title = "月度趋势"
    chart.y_axis.title = "金额"
    chart.width = 20
    chart.height = 12
    
    data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(monthly))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(monthly))
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A" + str(len(monthly) + 6))
    
    auto_adjust_column_width(ws)
    
    # 如果有分类列，创建分类汇总
    if category_col and category_col in df.columns:
        ws_category = wb.create_sheet('分类汇总')
        
        apply_title_style(ws_category, 1, 1, '分类汇总报表')
        
        category_stats = df.groupby(category_col).agg(
            金额合计=(amount_col, 'sum'),
            占比=(amount_col, lambda x: x.sum() / df[amount_col].sum())
        ).reset_index()
        
        category_stats = category_stats.sort_values('金额合计', ascending=False)
        
        headers = [category_col, '金额合计', '占比']
        for col_idx, header in enumerate(headers, 1):
            ws_category.cell(row=3, column=col_idx, value=header)
            apply_header_style(ws_category, 3, col_idx)
        
        for row_idx, (_, row) in enumerate(category_stats.iterrows(), 4):
            ws_category.cell(row=row_idx, column=1, value=row[category_col])
            ws_category.cell(row=row_idx, column=2, value=row['金额合计'])
            ws_category.cell(row=row_idx, column=3, value=row['占比'])
            
            for col_idx in range(1, 4):
                apply_data_style(ws_category, row_idx, col_idx, col_idx > 1)
            
            ws_category.cell(row=row_idx, column=2).number_format = NUMBER_FORMAT
            ws_category.cell(row=row_idx, column=3).number_format = PERCENTAGE_FORMAT
        
        auto_adjust_column_width(ws_category)
    
    # 保存到BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer
